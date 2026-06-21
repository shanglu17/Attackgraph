"""
Convert the real ZG-ONE source workbook into the v4 importable template format.

Source: 零重力/网络安保资产_边界及系统_接口和数据流清单20260616(1).xlsx
Output: docs/资产清单_v4_真实数据.xlsx  (same sheet names/headers/order as 资产清单_v4_可导入.xlsx)

Rules (confirmed with user):
- Operational systems (指挥控制系统/大数据中心 — 不在审查范围) are EXCLUDED.
- Sheets absent from the source (域属性表/信任边界表/威胁主体表/STRIDE映射表/数据资产)
  reuse the standard v4 reference content.
- Derived: BI ids (BI01..), BDF ids (BDF01..), SDF ids (SDF01..); 安保边界(SB) from 接口类别;
  blank 类型 rows forward-filled (continuation rows inherit the type above); 目标功能 = F-numbers
  extracted from 关联功能; BDF→BI link by matching producer/consumer to BI 外部实体/接入对象.
"""
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import sys
SRC = sys.argv[1] if len(sys.argv) > 1 else \
    r"D:/program download/weixin/xwechat_files/wxid_k5k44kfijop622_8d83/msg/file/2026-06/网络安保资产_边界及系统_接口和数据流清单_0618(1).xlsx"
OUT = r"E:/document/airness/project/Attackgraph/docs/资产清单_v4_真实数据.xlsx"
STOP = "以下是样例，请勿填写此行以下内容"

# ── styling helpers (mirror gen_template_v3.py) ──────────────────────────────
thin = Side(border_style="thin", color="CCCCCC")
def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def fill(c): return PatternFill("solid", fgColor=c)

def apply_hdr(ws, headers, widths, color):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = fill(color)
        c.alignment = ctr()
        c.border = bdr()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

def add_rows(ws, data, nc):
    for ri, row in enumerate(data, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)
        bg = "FFEBF4FF" if ri % 2 == 0 else "FFFFFFFF"
        for col in range(1, nc + 1):
            c = ws.cell(row=ri, column=col)
            c.font = Font(size=10)
            c.alignment = lft()
            c.border = bdr()
            c.fill = fill(bg)

def freeze(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ── source readers ──────────────────────────────────────────────────────────
src = load_workbook(SRC, data_only=True)

def sheet_rows(name):
    ws = src[name]
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if c is None else str(c).strip() for c in row])
    return out

def sheet_rows_expanded(name, expand_cols):
    """Like sheet_rows, but fills merged-cell ranges for the given 0-based columns.
    openpyxl returns a value only in a merged range's top-left cell; the source uses
    merged cells in 关联功能 (and others), so continuation rows read blank. We expand
    ONLY the requested columns (not 内容/描述) to avoid over-generating rows."""
    ws = src[name]
    grid = [["" if c is None else str(c).strip() for c in row]
            for row in ws.iter_rows(values_only=True)]
    want = {c + 1 for c in expand_cols}  # openpyxl columns are 1-based
    for mr in ws.merged_cells.ranges:
        cols = [c for c in range(mr.min_col, mr.max_col + 1) if c in want]
        if not cols:
            continue
        tl = grid[mr.min_row - 1][mr.min_col - 1]
        for r in range(mr.min_row - 1, mr.max_row):
            if r >= len(grid):
                continue
            for c1 in cols:
                if c1 - 1 < len(grid[r]):
                    grid[r][c1 - 1] = tl
    return grid

def ffill(rows, cols):
    last = {}
    out = []
    for r in rows:
        r = list(r) + [""] * 12
        for c in cols:
            if r[c].strip():
                last[c] = r[c]
            else:
                r[c] = last.get(c, "")
        out.append(r)
    return out

def fnums(s):
    return ",".join(dict.fromkeys(re.findall(r"F\d+", s or "")))

def norm(s):
    s = re.sub(r"[（(].*?[)）]", "", s or "")
    s = re.sub(r"\s+", "", s)
    return s.strip().upper()

# Synonyms between the 边界数据流 and 边界接口 sheets (same entity, different wording).
ALIAS = {
    "环境视觉目标": "开放环境", "环境激光反射目标": "开放环境",
    "环境音频源": "开放环境", "环境视频目标": "开放环境", "环境目标": "开放环境",
    "ICCP": "IMS", "ICC": "IMS", "BICC": "IMS",
    "RTK": "RTK服务",
}
_ALIAS_N = {norm(k): norm(v) for k, v in ALIAS.items()}

def canon(s):
    """Normalized form with synonym resolution — used only for BI matching, not for output names."""
    u = norm(s)
    return _ALIAS_N.get(u, u)

def clean_name(s):
    s = re.sub(r"[（(].*?[)）]", "", s or "")
    s = re.split(r"[、,，/\n\r]+", s)[0]
    return s.strip()

def split_multi(s):
    return [canon(p) for p in re.split(r"[、,，/\n\r]+", s or "") if canon(p)]

# ── Sheet: 边界接口表 (from 边界接口) ─────────────────────────────────────────
SB_BY_CAT = {
    "网络通信": "SB-01", "人机交互": "SB-01",
    "导航输入": "SB-02", "感知输入": "SB-02",
    "维护接入": "SB-03", "调试输入": "SB-03", "调试接入": "SB-03", "外围设备接口": "SB-03",
}
bi_raw = [r for r in sheet_rows("边界接口")[1:] if any(c.strip() for c in r)]
# 接口类别(0)/外部实体(1)/边界接入对象(2) 在源表里都用合并单元格,延续行为空,需向前填充。
bi_raw = ffill(bi_raw, [0, 1, 2])
bi_list = []          # rows for the sheet
bi_lookup = []        # (bi_id, ext_norm, [access_norms])
bi_n = 0
for r in bi_raw:
    cat, ext, access, phys, logic, dirn, desc = r[0], r[1], r[2], r[4], r[5], r[6], r[7]
    sb = SB_BY_CAT.get(cat, "SB-03")
    # A BI whose 边界接入对象 lists several subsystems (e.g. "IMS、FMS、FCS") is split into
    # one single-access BI per subsystem, so each has an unambiguous FP entry.
    parts = [p.strip() for p in re.split(r"[、,，/\n\r]+", access) if p.strip()]
    if len(parts) <= 1:
        parts = [access]  # keep original (may be empty)
    for part in parts:
        bi_n += 1
        bid = f"BI{bi_n:02d}"
        bi_list.append([bid, cat, ext, part, phys, logic, dirn, sb, desc])
        bi_lookup.append((bid, canon(ext), split_multi(part)))

def _sub(a, b):
    """True if normalized a and b overlap (either is a substring of the other)."""
    return bool(a) and bool(b) and (a in b or b in a)

def match_bi(prod, cons, dest):
    p, c, d = canon(prod), canon(cons), canon(dest)
    onboard = [x for x in (c, d) if x]
    for bid, ext, accs in bi_lookup:                       # exact directional
        if ext and p == ext and any(a == x for a in accs for x in onboard):
            return bid
        if ext and (c == ext or d == ext) and p in accs:
            return bid
    for bid, ext, accs in bi_lookup:                       # two-way substring fallback
        if ext and _sub(ext, p) and any(_sub(a, x) for a in accs for x in onboard):
            return bid
        if ext and (_sub(ext, c) or _sub(ext, d)) and any(_sub(a, p) for a in accs):
            return bid
    return ""

# ── Sheet: 边界数据流表 (from 边界数据流) ─────────────────────────────────────
OOS = {norm(x) for x in ("指挥控制系统", "大数据中心", "大数据平台")}
# 关联功能(col 5) uses merged cells in the source → expand it, else continuation rows read blank.
bdf_raw = [r for r in sheet_rows_expanded("边界数据流", [5])[1:] if any(c.strip() for c in r)]
bdf_raw = ffill(bdf_raw, [0, 1, 2, 4])  # Producer, Consumer, Destination, 类型
bdf_list = []
unmatched = 0
n = 0
for r in bdf_raw:
    prod, cons, dest, desc, typ, func = r[0], r[1], r[2], r[3], r[4], r[5]
    if not desc.strip():
        continue
    if norm(prod) in OOS or norm(cons) in OOS:
        continue
    fcodes = fnums(func)
    if not fcodes and "人工控制输入" in desc:
        fcodes = "F7"  # 操控员 manual control input = HMI interaction (F7), per review
    # A multi-subsystem consumer ("IMS、FMS、FCS") becomes one BDF per subsystem.
    cons_parts = [p.strip() for p in re.split(r"[、,，/\n\r]+", cons) if p.strip()]
    if len(cons_parts) <= 1:
        cons_parts = [cons]
    for cpart in cons_parts:
        n += 1
        bid = match_bi(prod, cpart, dest)
        if not bid:
            unmatched += 1
        bdf_list.append([f"BDF{n:02d}", clean_name(prod), clean_name(cpart), desc, typ.upper(), fcodes, bid])

# ── Sheet: 系统数据流表 (from 系统间数据流) ───────────────────────────────────
sdf_raw = [r for r in sheet_rows_expanded("系统间数据流", [5])[1:] if any(c.strip() for c in r)]
sdf_raw = ffill(sdf_raw, [0, 1, 4])  # Producer, Consumer, 类型
sdf_list = []
m = 0
for r in sdf_raw:
    prod, cons, content, typ, func = r[0], r[1], r[3], r[4], r[5]
    if not content.strip():
        continue
    m += 1
    sdf_list.append([f"SDF{m:02d}", clean_name(prod), clean_name(cons), content, typ.upper(), fnums(func)])

# ── consolidate STATE / DATA flows sharing (producer, consumer, type) into one ──
# CMD/CONFIG/LOAD/ALERT/SENSOR are kept separate (per review); only STATE & DATA merge.
MERGE_TYPES = {"STATE", "DATA"}

def consolidate(rows, prod_i, cons_i, type_i, desc_i, func_i):
    groups = {}
    order = []
    for row in rows:
        typ = row[type_i].upper()
        key = (row[prod_i], row[cons_i], typ) if typ in MERGE_TYPES else ("__solo__", id(row))
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)
    out = []
    for key in order:
        members = groups[key]
        base = list(members[0])
        if len(members) > 1:
            descs = []
            for mm in members:
                d = mm[desc_i].strip()
                if d and d not in descs:
                    descs.append(d)
            base[desc_i] = "\n".join(descs)
            funcs = []
            for mm in members:
                for f in re.findall(r"F\d+", mm[func_i] or ""):
                    if f not in funcs:
                        funcs.append(f)
            base[func_i] = ",".join(sorted(funcs, key=lambda x: int(x[1:])))
        out.append(base)
    return out

bdf_list = consolidate(bdf_list, 1, 2, 4, 3, 5)
for i, row in enumerate(bdf_list, 1):
    row[0] = f"BDF{i:02d}"
sdf_list = consolidate(sdf_list, 1, 2, 4, 3, 5)
for i, row in enumerate(sdf_list, 1):
    row[0] = f"SDF{i:02d}"

# ── Sheet: 功能资产 (from 系统功能) ───────────────────────────────────────────
CRIT_BY_FN = {"安全关键": "High", "安全相关": "Medium", "运行支撑": "Low"}
fn_rows = [r for r in sheet_rows("系统功能")[1:] if (r[0] or "").strip().upper().startswith("F")]
func_list = []
for r in fn_rows:
    fid, name, desc, crit = r[0].strip(), r[1], r[2], r[3]
    func_list.append([f"SF.{fid}", name, f"{fid}: {desc}（关键性:{crit}）"])

# ── Sheet: 支持资产 (subsystems from BDF + SDF producers/consumers) ───────────
ONBOARD_CORE = {"ICP", "ICCP", "IMS", "EPS", "ES", "PACKS", "HVDS", "NAVS", "DAAS", "PES", "FCS", "FMS", "RA"}
def classify(name):
    u = norm(name)
    if u in ONBOARD_CORE:
        return ("Internal", "High")
    if u == "DLS":
        return ("Shared", "High")
    if u == "AVCS":
        return ("Shared", "Medium")
    return ("External", "Medium")

entities = []
seen = set()
for src_list, ia, ib in ((bdf_list, 1, 2), (sdf_list, 1, 2)):
    for row in src_list:
        for name in (row[ia], row[ib]):
            key = norm(name)
            if name.strip() and key and key not in seen and key not in OOS:
                seen.add(key)
                entities.append(name.strip())
support_list = []
for i, name in enumerate(entities, 1):
    dom, crit = classify(name)
    support_list.append([f"ASA.{i:02d}", name, "", dom, crit])

# ── reused standard reference sheets (from gen_template_v3.py) ────────────────
rows_da = [
    ["DA.01", "ICC_APP", "外场可加载软件", "LOAD", "", "D.6", "综合控制计算机(ICC/IMS)主应用软件", "F1,F2,F3"],
    ["DA.02", "BICC_APP", "外场可加载软件", "LOAD", "", "D.6", "备份综合控制计算机(BICC)应用软件", "F1,F2,F3"],
    ["DA.03", "MCU_APP", "外场可加载软件", "LOAD", "", "D.6", "电机控制器(MCU/EPS)管理软件", "F1,F2"],
    ["DA.04", "BMS_APP", "外场可加载软件", "LOAD", "", "D.6", "电池管理系统(BMS/PACKS)全功能软件", "F1,F3,F9"],
    ["DA.05", "NAVS_FW", "外场可加载软件", "LOAD", "", "D.6", "组合导航系统固件（光纤/MEMS）", "F2,F3,F4"],
    ["DA.06", "DAAS_CONFIG", "可加载配置文件", "CONFIG", "", "D.6", "探测与避让系统（DAA）配置参数", "F2,F3,F13"],
    ["DA.07", "ICC_PARAM", "可加载配置文件", "CONFIG", "", "D.6", "ICC飞行参数/导航参数/系统配置", "F1,F2,F3"],
    ["DA.08", "MCU_PARAM", "可加载配置文件", "CONFIG", "", "D.6", "电机控制器配置参数", "F1,F2"],
]
rows_dm = [
    ["D.1", "外部非信任域", "Untrusted", "External", "RCS遥控站/转发服务器/GNSS/RTK/地面维护设备/大数据平台等外部实体。来自此域的数据不默认可信。"],
    ["D.2", "空地数据链", "Untrusted", "External", "DLS与地面RCS间的射频通信链路（1.4G/4G/5G）。可能经过公网/运营商网络。"],
    ["D.3", "机载边界域", "Semi-Trusted", "Shared", "DLS/AVCS等机载边界与通信系统，承担对外音视频业务及链路接入功能。"],
    ["D.4", "机载航电域", "Trusted", "Internal", "IMS/DAAS/NAVS等航空电子设备。部分依赖开放环境输入（GNSS/DAA），需一致性校验。"],
    ["D.5", "机载核心信任域", "Trusted", "Internal", "IMS(ICC/BICC)/FCS/FMS/EPS/ES/PACKS/HVDS/PES。直接决定飞行安全，内部通信假定在受控环境。"],
    ["D.6", "地面维护域", "Semi-Trusted", "External", "地面维护设备接入接口（CAN/ETH/RS422/USB等）。具备软件加载/参数配置/诊断权限，属高权限安保入口。"],
]
rows_tb = [
    ["SB-01", "外部网络与通信边界", "系统与外部网络、云平台、通信基础设施及操控终端之间的数据交换边界。", "D.1;D.2;D.3", "TA-E-01;TA-E-02;TA-I-01", "是"],
    ["SB-02", "开放环境输入边界", "系统接收来自开放环境的导航、感知和探测信息的边界。", "D.1;D.4", "TA-E-03;TA-E-04", "是"],
    ["SB-03", "维护与物理接入边界", "系统软件加载、参数配置、调试维护和设备接入边界。", "D.6;D.5", "TA-E-05;TA-I-02;TA-I-03;TA-I-04", "是"],
]
rows_ta = [
    ["TA-E-01", "远程网络攻击者", "外部", "经公网/运营商网络对空地数据链及云平台发起远程攻击。"],
    ["TA-E-02", "通信链路攻击者", "外部", "针对1.4G/4G/5G射频链路的中间人/重放/注入攻击。"],
    ["TA-E-03", "GNSS欺骗/干扰者", "外部", "对GNSS卫星信号进行欺骗或干扰，影响导航定位。"],
    ["TA-E-04", "开放环境输入伪造者", "外部", "伪造探测/感知/导航等开放环境输入数据。"],
    ["TA-E-05", "恶意维护人员/设备", "外部", "通过维护接口加载恶意软件或篡改参数。"],
    ["TA-I-01", "受感染地面终端", "内部", "被攻陷的RCS/转发服务器成为内部攻击跳板。"],
    ["TA-I-02", "维护误操作", "内部", "维护人员误操作导致配置或软件异常。"],
    ["TA-I-03", "供应链植入", "第三方", "供应链环节植入的后门或被篡改的组件。"],
    ["TA-I-04", "内部越权访问", "内部", "内部人员越权访问或滥用维护/管理权限。"],
]
stride_data = [
    ["CMD", "Spoofing + Tampering", "未授权指令注入/篡改/重放；推力/上下电/伞降误触发", "双向认证 + 完整性保护(MAC/签名) + 序列号/时间窗 + 运行状态门控", "Catastrophic / Hazardous"],
    ["CONFIG", "Tampering", "配置数据恶意修改(航线/围栏)；未授权维护参数篡改", "签名 + 语义约束检查 + 版本控制 + 强认证/RBAC", "Hazardous / Major"],
    ["LOAD", "Tampering", "恶意软件加载(维护接口)；未授权维护访问", "Secure Boot + 软件签名验证 + 完整性链 + 强认证/RBAC", "Catastrophic"],
    ["STATE", "Spoofing", "状态/遥测伪造；导航融合异常", "多源融合一致性校验 + 异常检测 + 冗余", "Major / Hazardous"],
    ["ALERT", "Tampering + DenialOfService", "告警抑制/篡改；总线阻塞DoS", "完整性保护 + 实时调度/带宽隔离 + 告警超时检测", "Hazardous / Catastrophic"],
    ["DATA", "InformationDisclosure", "状态/视频数据泄露；日志导出暴露", "加密传输(TLS) + 访问控制 + 流媒体加密", "Minor / Major"],
    ["SENSOR", "Spoofing + Tampering", "GNSS欺骗/干扰；探测/感知输入伪造", "多源一致性 + 物理合理性约束 + 冗余传感器交叉验证", "Major / Hazardous"],
]

# ── build workbook ───────────────────────────────────────────────────────────
wb = Workbook()

def make(title, headers, widths, color, data, dv=None):
    ws = wb.create_sheet(title)
    apply_hdr(ws, headers, widths, color)
    if dv:
        col, formula = dv
        v = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(v)
        v.sqref = f"{col}2:{col}500"
    add_rows(ws, data + [[STOP] + [""] * (len(headers) - 1)], len(headers))
    freeze(ws)
    return ws

ws1 = make("边界数据流表",
           ["边界数据流编号", "产生者", "用户", "数据流描述", "数据流类型", "目标功能", "所属边界接口"],
           [14, 16, 16, 44, 12, 16, 12], "FF2B6CB0", bdf_list, dv=("E", '"CMD,CONFIG,STATE,DATA,LOAD,ALERT,SENSOR"'))
wb.remove(wb["Sheet"])  # drop default
make("边界接口表",
     ["接口编号", "接口类别", "外部实体", "边界接入对象", "物理互连类型", "逻辑协议", "方向性", "安保边界", "接口说明"],
     [10, 14, 18, 18, 14, 14, 10, 12, 30], "FF1D4ED8", bi_list)
make("数据资产",
     ["编号", "数据名称", "数据类型", "数据流类型", "对应接口", "所属域", "说明", "目标功能"],
     [10, 26, 22, 12, 16, 10, 36, 14], "FF2F855A", rows_da, dv=("D", '"CMD,CONFIG,STATE,DATA,LOAD,ALERT"'))
make("域属性表",
     ["域编号", "域名称", "信任程度", "安全域", "描述"],
     [10, 22, 14, 14, 52], "FF6B46C1", rows_dm)
make("功能资产",
     ["编号", "功能资产名称", "资产说明"],
     [10, 24, 60], "FF975A16", func_list)
make("支持资产",
     ["编号", "名称", "交联接口", "安全域", "关键性"],
     [10, 22, 30, 12, 10], "FF975A16", support_list)
make("STRIDE映射表",
     ["数据流类型", "主要STRIDE类别", "典型攻击模式", "安保措施方向", "FHA严重度参考"],
     [14, 26, 42, 48, 20], "FF9B2C2C", stride_data)
make("信任边界表",
     ["边界编号", "边界名称", "边界说明", "覆盖域", "相关威胁主体", "是否进入内部传播"],
     [12, 24, 46, 18, 28, 16], "FF0E7490", rows_tb)
make("威胁主体表",
     ["主体编号", "名称", "类型", "说明"],
     [12, 24, 12, 50], "FFB91C1C", rows_ta)
make("系统数据流表",
     ["数据流编号", "产生者", "用户", "内容", "数据流类型", "目标功能"],
     [12, 12, 12, 50, 12, 16], "FF0F766E", sdf_list, dv=("E", '"CMD,CONFIG,STATE,DATA,LOAD,ALERT,SENSOR"'))

wb.save(OUT)
print("Saved:", OUT)
print(f"  边界接口表(BI)   : {len(bi_list)}")
print(f"  边界数据流表(BDF): {len(bdf_list)}  (未匹配到BI: {unmatched})")
print(f"  系统数据流表(SDF): {len(sdf_list)}")
print(f"  功能资产         : {len(func_list)}")
print(f"  支持资产         : {len(support_list)}")
