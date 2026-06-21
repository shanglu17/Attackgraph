"""
Generate docs/资产清单_v4_可导入.xlsx

Source of truth = ZG-ONE vol1 (BI/BDF) + vol3 第四章 (SB/TA/F).
边界数据流表 = vol1 表A-1 边界数据流 (BDF01-BDF81, 航空器↔外部, 含 Producer/Consumer/类型/功能 + 所属BI)
边界接口表 = vol1 表7-1 边界接口清单 (BI01-BI21, 含 安保边界SB)
信任边界表 / 威胁主体表 = vol3 表4-1 + vol1 表5-4
功能资产 F1-F13, 域属性表 D.1-D.6 = vol1/vol3
模型: SB(信任边界) → BI(边界接口) → BDF(边界数据流)
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
thin = Side(border_style="thin", color="CCCCCC")


def bdr():
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def bold_w():
    return Font(bold=True, size=10, color="FFFFFF")


def norm():
    return Font(size=10)


def ctr():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def lft():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def fill(c):
    return PatternFill("solid", fgColor=c)


def apply_hdr(ws, headers, widths, color):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold_w()
        c.fill = fill(color)
        c.alignment = ctr()
        c.border = bdr()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28


def style_row(ws, ri, nc, alt):
    bg = "FFEBF4FF" if alt else "FFFFFFFF"
    for col in range(1, nc + 1):
        c = ws.cell(row=ri, column=col)
        c.font = norm()
        c.alignment = lft()
        c.border = bdr()
        c.fill = fill(bg)


def add_rows(ws, data, nc):
    for ri, row in enumerate(data, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)
        style_row(ws, ri, nc, ri % 2 == 0)


def freeze(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def clean_funcs(raw):
    return ",".join(dict.fromkeys(re.findall(r"F\d+", raw)))


STOP = "以下是样例，请勿填写此行以下内容"

# ─────────────────────────────────────────────────────────────────────────────
# BI table (vol1 表7-1): (id, class, external_entity, access_object, physical, logical, direction, SB, desc)
# ─────────────────────────────────────────────────────────────────────────────
bi_rows = [
    ("BI01", "网络通信", "大数据平台", "RCS", "ETH", "HTTP", "双向", "SB-01", "云业务接口"),
    ("BI02", "网络通信", "转发服务器", "RCS", "ETH", "QUIC", "双向", "SB-01", "外部链路与业务中继接口"),
    ("BI03", "网络通信", "转发服务器", "DLS", "4G/5G", "QUIC", "双向", "SB-01", "空地链路中继接口"),
    ("BI04", "导航输入", "GNSS", "NAVS", "RF", "GNSS信号", "单向", "SB-02", "GNSS导航输入接口"),
    ("BI05", "导航输入", "RTK服务", "RCS", "RS422", "私有协议", "单向", "SB-02", "RTK差分修正输入接口"),
    ("BI06", "感知输入", "开放环境", "DAAS", "摄像头链路", "图像数据流", "单向", "SB-02", "视觉感知输入接口"),
    ("BI07", "感知输入", "开放环境", "DAAS", "激光雷达", "点云数据流", "单向", "SB-02", "激光雷达输入接口"),
    ("BI08", "感知输入", "开放环境", "AVCS", "麦克风", "音频数据流", "单向", "SB-02", "环境音频输入接口"),
    ("BI09", "感知输入", "开放环境", "AVCS", "摄像头", "视频数据流", "单向", "SB-02", "环境视频输入接口"),
    ("BI10", "维护接入", "地面维护设备", "ICP", "CAN/ETH", "私有协议", "双向", "SB-03", "ICP维护接口"),
    ("BI11", "维护接入", "地面维护设备", "DLS", "ETH", "私有协议", "双向", "SB-03", "DLS维护接口"),
    ("BI12", "维护接入", "地面维护设备", "EPS", "CAN", "CAN协议", "双向", "SB-03", "EPS维护接口"),
    ("BI13", "维护接入", "地面维护设备", "PACKS", "CAN", "CAN协议", "双向", "SB-03", "PACKS维护接口"),
    ("BI14", "维护接入", "地面维护设备", "HVDS", "CAN", "CAN协议", "双向", "SB-03", "HVDS维护接口"),
    ("BI15", "维护接入", "地面维护设备", "DAAS", "USB3.0", "私有协议", "双向", "SB-03", "DAAS维护接口"),
    ("BI16", "维护接入", "地面维护设备", "AVCS", "RS232/ETH", "私有协议", "双向", "SB-03", "AVCS维护接口"),
    ("BI17", "维护接入", "地面维护设备", "NAVS", "RS422", "私有协议", "双向", "SB-03", "NAVS维护接口"),
    ("BI18", "维护接入", "地面维护设备", "PES", "RS232", "私有协议", "双向", "SB-03", "PES维护接口"),
    ("BI19", "调试接入", "调试设备", "RCS", "RS422", "私有协议", "双向", "SB-03", "RCS调试接口"),
    ("BI20", "外围设备接口", "UPS", "RCS", "RS232", "私有协议", "双向", "SB-03", "电源监控接口"),
    ("BI21", "人机交互", "操控员", "RCS", "人机交互设备", "HMI协议", "双向", "SB-01", "操控员交互接口"),
]
bi_to_sb = {r[0]: r[7] for r in bi_rows}

# ─────────────────────────────────────────────────────────────────────────────
# BDF table (vol1 表A-1): (bdf, bi, producer, consumer, desc, type, func)
# ─────────────────────────────────────────────────────────────────────────────
bdf_rows = [
    ("BDF01", "BI01", "RCS", "大数据平台", "用户账号密码", "DATA", "F8"),
    ("BDF02", "BI01", "RCS", "大数据平台", "本地航线数据、电子围栏、备降点", "CONFIG", "F3/F4"),
    ("BDF03", "BI01", "RCS", "大数据平台", "飞行数据", "DATA", "F10"),
    ("BDF04", "BI01", "大数据平台", "RCS", "用户登录结果", "DATA", "F8"),
    ("BDF05", "BI01", "大数据平台", "RCS", "航线、电子围栏、备降点数据", "CONFIG", "F3/F4"),
    ("BDF06", "BI01", "大数据平台", "RCS", "航空器在线状态信息", "STATE", "F7/F8"),
    ("BDF07", "BI01", "大数据平台", "RCS", "气象信息", "DATA", "F3/F4"),
    ("BDF08", "BI02", "RCS", "转发服务器", "链路切换指令", "CMD", "F8"),
    ("BDF09", "BI02", "RCS", "转发服务器", "遥控站状态", "STATE", "F8"),
    ("BDF10", "BI02", "RCS", "转发服务器", "航线、电子围栏、备降数据", "CONFIG", "F3/F4"),
    ("BDF11", "BI02", "RCS", "转发服务器", "音频数据", "DATA", "F8"),
    ("BDF12", "BI02", "RCS", "转发服务器", "飞行控制类指令", "CMD", "F2/F8"),
    ("BDF13", "BI02", "RCS", "转发服务器", "绑定航空器指令", "CMD", "F8"),
    ("BDF14", "BI02", "转发服务器", "RCS", "ACU状态数据", "STATE", "F8"),
    ("BDF15", "BI02", "转发服务器", "RCS", "航空器综合管理系统状态", "STATE", "F7"),
    ("BDF16", "BI02", "转发服务器", "RCS", "航空器飞行管理系统状态", "STATE", "F3/F4/F7"),
    ("BDF17", "BI02", "转发服务器", "RCS", "航空器飞行控制系统状态", "STATE", "F2/F7"),
    ("BDF18", "BI02", "转发服务器", "RCS", "航空器电推进系统状态", "STATE", "F1/F7"),
    ("BDF19", "BI02", "转发服务器", "RCS", "航空器动力电池系统状态", "STATE", "F1/F7"),
    ("BDF20", "BI02", "转发服务器", "RCS", "航空器高压配电系统状态", "STATE", "F1/F7"),
    ("BDF21", "BI02", "转发服务器", "RCS", "航空器电气系统状态", "STATE", "F7/F9"),
    ("BDF22", "BI02", "转发服务器", "RCS", "航空器伞降系统状态", "STATE", "F5/F7"),
    ("BDF23", "BI02", "转发服务器", "RCS", "航空器导航系统状态", "STATE", "F4/F7"),
    ("BDF24", "BI02", "转发服务器", "RCS", "航空器探测与避让系统状态", "STATE", "F7/F13"),
    ("BDF25", "BI02", "转发服务器", "RCS", "航空器音视频通信系统状态", "STATE", "F6/F7"),
    ("BDF26", "BI02", "转发服务器", "RCS", "航空器存储的航线数据", "CONFIG", "F3/F4"),
    ("BDF27", "BI02", "转发服务器", "RCS", "音视频数据", "DATA", "F6/F8"),
    ("BDF28", "BI03", "转发服务器", "DLS", "链路切换指令", "CMD", "F8"),
    ("BDF29", "BI03", "转发服务器", "DLS", "遥控站状态", "STATE", "F8"),
    ("BDF30", "BI03", "转发服务器", "DLS", "转发服务器状态", "STATE", "F8"),
    ("BDF31", "BI03", "转发服务器", "DLS", "航线、电子围栏、备降数据", "CONFIG", "F3/F4"),
    ("BDF32", "BI03", "转发服务器", "DLS", "音频数据", "DATA", "F8"),
    ("BDF33", "BI03", "转发服务器", "DLS", "飞行控制类指令", "CMD", "F2/F8"),
    ("BDF34", "BI03", "转发服务器", "DLS", "绑定航空器指令", "CMD", "F8"),
    ("BDF35", "BI03", "DLS", "转发服务器", "ACU状态数据", "STATE", "F8"),
    ("BDF36", "BI03", "DLS", "转发服务器", "航空器综合管理系统状态", "STATE", "F7"),
    ("BDF37", "BI03", "DLS", "转发服务器", "航空器飞行管理系统状态", "STATE", "F3/F4/F7"),
    ("BDF38", "BI03", "DLS", "转发服务器", "航空器飞行控制系统状态", "STATE", "F2/F7"),
    ("BDF39", "BI03", "DLS", "转发服务器", "航空器电推进系统状态", "STATE", "F1/F7"),
    ("BDF40", "BI03", "DLS", "转发服务器", "航空器动力电池系统状态", "STATE", "F1/F7"),
    ("BDF41", "BI03", "DLS", "转发服务器", "航空器高压配电系统状态", "STATE", "F1/F7"),
    ("BDF42", "BI03", "DLS", "转发服务器", "航空器电气系统状态", "STATE", "F7/F9"),
    ("BDF43", "BI03", "DLS", "转发服务器", "航空器伞降系统状态", "STATE", "F5/F7"),
    ("BDF44", "BI03", "DLS", "转发服务器", "航空器导航系统状态", "STATE", "F4/F7"),
    ("BDF45", "BI03", "DLS", "转发服务器", "航空器探测与避让系统状态", "STATE", "F7/F13"),
    ("BDF46", "BI03", "DLS", "转发服务器", "航空器音视频通信系统状态", "STATE", "F6/F7"),
    ("BDF47", "BI03", "DLS", "转发服务器", "航空器存储的航线数据", "CONFIG", "F3/F4"),
    ("BDF48", "BI03", "DLS", "转发服务器", "音视频数据", "DATA", "F6/F8"),
    ("BDF49", "BI04", "GNSS", "NAVS", "GNSS定位与时间信息", "SENSOR", "F4"),
    ("BDF50", "BI05", "RTK", "RCS", "RTK差分修正数据", "SENSOR", "F4"),
    ("BDF51", "BI06", "环境视觉目标", "DAAS", "环境视觉图像数据", "SENSOR", "F13"),
    ("BDF52", "BI07", "环境激光反射目标", "DAAS", "激光雷达点云数据", "SENSOR", "F13"),
    ("BDF53", "BI08", "环境音频源", "AVCS", "环境语音与音频输入数据、机舱环境音频数据", "DATA", "F6/F8"),
    ("BDF54", "BI09", "环境视频目标", "AVCS", "舱内视频采集数据", "DATA", "F6/F8"),
    ("BDF55", "BI09", "环境视频目标", "AVCS", "下视环境视频数据", "DATA", "F6/F8"),
    ("BDF56", "BI10", "地面维护设备", "ICP", "软件加载数据", "LOAD", "F11"),
    ("BDF57", "BI10", "地面维护设备", "ICP", "参数配置数据", "CONFIG", "F2/F3/F4/F11/F12"),
    ("BDF58", "BI10", "ICP", "地面维护设备", "诊断数据、日志数据", "DATA", "F10/F11"),
    ("BDF59", "BI10", "ICP", "地面维护设备", "软件版本与配置状态", "STATE", "F11"),
    ("BDF60", "BI11", "地面维护设备", "DLS", "配置文件", "CONFIG", "F8"),
    ("BDF61", "BI11", "地面维护设备", "DLS", "管理软件", "LOAD", "F8"),
    ("BDF62", "BI11", "DLS", "地面维护设备", "日志", "DATA", "F10"),
    ("BDF63", "BI12", "地面维护设备", "EPS", "配置文件", "CONFIG", "F1"),
    ("BDF64", "BI12", "地面维护设备", "EPS", "管理软件/引导加载软件", "LOAD", "F1"),
    ("BDF65", "BI13", "地面维护设备", "PACKS", "管理软件/引导加载软件", "LOAD", "F1/F9"),
    ("BDF66", "BI14", "地面维护设备", "HVDS", "管理软件/引导加载软件", "LOAD", "F1/F9"),
    ("BDF67", "BI15", "地面维护设备", "DAAS", "配置文件", "CONFIG", "F11/F13"),
    ("BDF68", "BI15", "DAAS", "地面维护设备", "日志", "DATA", "F10"),
    ("BDF69", "BI16", "地面维护设备", "AVCS", "配置/诊断数据", "CONFIG", "F7/F8"),
    ("BDF70", "BI16", "地面维护设备", "AVCS", "管理软件/引导加载软件", "LOAD", "F6/F8"),
    ("BDF71", "BI16", "AVCS", "地面维护设备", "AVCS运行数据备份", "DATA", "F10"),
    ("BDF72", "BI17", "地面维护设备", "NAVS", "管理软件/引导加载软件", "LOAD", "F4"),
    ("BDF73", "BI17", "地面维护设备", "NAVS", "软件维护配置", "CONFIG", "F4"),
    ("BDF74", "BI18", "PES", "地面维护设备", "日志", "DATA", "F5/F10"),
    ("BDF75", "BI18", "地面维护设备", "PES", "引导加载软件", "LOAD", "F5/F11"),
    ("BDF76", "BI19", "调试计算机", "RCS", "调试控制数据", "CMD", "F2/F11"),
    ("BDF77", "BI20", "UPS", "RCS", "UPS运行状态", "STATE", "F9"),
    ("BDF78", "BI21", "操控员", "RCS", "人工控制输入", "CMD", "F2/F8"),
    ("BDF79", "BI21", "RCS", "操控员", "航空器运行状态显示信息", "STATE", "F7/F8"),
    ("BDF80", "BI21", "RCS", "操控员", "告警显示信息", "ALERT", "F5/F7/F12"),
    ("BDF81", "BI21", "RCS", "操控员", "音视频与任务显示信息", "DATA", "F6/F7/F8"),
]

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: 边界数据流表  ←  BDF (边界数据流, 航空器↔外部)
# cols: 边界数据流编号(BDF) | 产生者 | 用户 | 数据流描述 | 数据流类型 | 目标功能 | 所属边界接口(BI)
# (物理/逻辑接口在「边界接口表」, SB 由 BI 自动推出, 故不在此重复)
# ─────────────────────────────────────────────────────────────────────────────
IF_H = ["边界数据流编号", "产生者", "用户", "数据流描述", "数据流类型", "目标功能", "所属边界接口"]
IF_W = [14, 14, 14, 42, 12, 18, 12]

ws1 = wb.active
ws1.title = "边界数据流表"
apply_hdr(ws1, IF_H, IF_W, "FF2B6CB0")

dv1 = DataValidation(type="list",
                     formula1='"CMD,CONFIG,STATE,DATA,LOAD,ALERT,SENSOR"',
                     allow_blank=True, showDropDown=False)
ws1.add_data_validation(dv1)
dv1.sqref = "E2:E500"

rows_if = []
for bdf, bi, prod, cons, content, dft, rawf in bdf_rows:
    # [边界数据流编号=BDF, 产生者, 用户, 描述, 类型, 目标功能, 所属边界接口BI]
    rows_if.append([bdf, prod, cons, content, dft, clean_funcs(rawf), bi])
rows_if.append([STOP] + [""] * (len(IF_H) - 1))

add_rows(ws1, rows_if, len(IF_H))
freeze(ws1)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: 边界接口表  ←  vol1 表7-1 (BI01-BI21, 安保边界SB)
# ─────────────────────────────────────────────────────────────────────────────
ws_bi = wb.create_sheet("边界接口表")
BI_H = ["接口编号", "接口类别", "外部实体", "边界接入对象", "物理互连类型", "逻辑协议", "方向性", "安保边界", "接口说明"]
BI_W = [10, 14, 16, 16, 16, 16, 10, 12, 24]
apply_hdr(ws_bi, BI_H, BI_W, "FF1D4ED8")

rows_bi = [list(r) for r in bi_rows]
rows_bi.append([STOP] + [""] * (len(BI_H) - 1))
add_rows(ws_bi, rows_bi, len(BI_H))
freeze(ws_bi)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: 数据资产  (ZG-ONE 关键可加载软件/配置 — 补充资产)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("数据资产")
DA_H = ["编号", "数据名称", "数据类型", "数据流类型", "对应接口", "所属域", "说明", "目标功能"]
DA_W = [10, 26, 22, 12, 16, 10, 36, 14]
apply_hdr(ws2, DA_H, DA_W, "FF2F855A")

dv2 = DataValidation(type="list",
                     formula1='"CMD,CONFIG,STATE,DATA,LOAD,ALERT"',
                     allow_blank=True, showDropDown=False)
ws2.add_data_validation(dv2)
dv2.sqref = "D2:D200"

rows_da = [
    ["DA.01", "ICC_APP",     "外场可加载软件", "LOAD",   "", "D.6", "综合控制计算机(ICC/IMS)主应用软件", "F1,F2,F3"],
    ["DA.02", "BICC_APP",    "外场可加载软件", "LOAD",   "", "D.6", "备份综合控制计算机(BICC)应用软件",  "F1,F2,F3"],
    ["DA.03", "MCU_APP",     "外场可加载软件", "LOAD",   "", "D.6", "电机控制器(MCU/EPS)管理软件",       "F1,F2"],
    ["DA.04", "BMS_APP",     "外场可加载软件", "LOAD",   "", "D.6", "电池管理系统(BMS/PACKS)全功能软件", "F1,F3,F9"],
    ["DA.05", "NAVS_FW",     "外场可加载软件", "LOAD",   "", "D.6", "组合导航系统固件（光纤/MEMS）",     "F2,F3,F4"],
    ["DA.06", "DAAS_CONFIG", "可加载配置文件", "CONFIG", "", "D.6", "探测与避让系统（DAA）配置参数",     "F2,F3,F13"],
    ["DA.07", "ICC_PARAM",   "可加载配置文件", "CONFIG", "", "D.6", "ICC飞行参数/导航参数/系统配置",     "F1,F2,F3"],
    ["DA.08", "MCU_PARAM",   "可加载配置文件", "CONFIG", "", "D.6", "电机控制器配置参数",                "F1,F2"],
    [STOP, "", "", "", "", "", "", ""],
]
add_rows(ws2, rows_da, len(DA_H))
freeze(ws2)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: 域属性表  (vol3 §4.2 trust domains)
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("域属性表")
DM_H = ["域编号", "域名称", "信任程度", "安全域", "描述"]
DM_W = [10, 22, 14, 14, 52]
apply_hdr(ws3, DM_H, DM_W, "FF6B46C1")

dv3a = DataValidation(type="list", formula1='"Trusted,Semi-Trusted,Untrusted"', allow_blank=False)
ws3.add_data_validation(dv3a)
dv3a.sqref = "C2:C100"
dv3b = DataValidation(type="list", formula1='"Internal,External,DMZ,Shared"', allow_blank=False)
ws3.add_data_validation(dv3b)
dv3b.sqref = "D2:D100"

rows_dm = [
    ["D.1", "外部非信任域",   "Untrusted",    "External",  "RCS遥控站/转发服务器/GNSS/RTK/地面维护设备/大数据平台等外部实体。来自此域的数据不默认可信。"],
    ["D.2", "空地数据链",     "Untrusted",    "External",  "DLS与地面RCS间的射频通信链路（1.4G/4G/5G）。可能经过公网/运营商网络。"],
    ["D.3", "机载边界域",     "Semi-Trusted", "Shared",    "DLS/AVCS等机载边界与通信系统，承担对外音视频业务及链路接入功能。"],
    ["D.4", "机载航电域",     "Trusted",      "Internal",  "IMS/DAAS/NAVS等航空电子设备。部分依赖开放环境输入（GNSS/DAA），需一致性校验。"],
    ["D.5", "机载核心信任域", "Trusted",      "Internal",  "IMS(ICC/BICC)/FCS/FMS/EPS/ES/PACKS/HVDS/PES。直接决定飞行安全，内部通信假定在受控环境。"],
    ["D.6", "地面维护域",     "Semi-Trusted", "External",  "地面维护设备接入接口（CAN/ETH/RS422/USB等）。具备软件加载/参数配置/诊断权限，属高权限安保入口。"],
    [STOP, "", "", "", ""],
]
add_rows(ws3, rows_dm, len(DM_H))
freeze(ws3)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: 功能资产  (vol1/vol3 F1-F13)
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("功能资产")
FN_H = ["编号", "功能资产名称", "资产说明"]
FN_W = [8, 24, 54]
apply_hdr(ws4, FN_H, FN_W, "FF975A16")

rows_fn = [
    ["SF.F1",  "提供飞行动力",       "F1: 动力推进控制（EPS/MCU）"],
    ["SF.F2",  "提供飞行控制",       "F2: 姿态控制与飞行执行（FCS/MCU）"],
    ["SF.F3",  "提供飞行管理功能",   "F3: 航迹规划/导航管理/任务管理（FMS）"],
    ["SF.F4",  "提供导航功能",       "F4: 多源导航融合/GNSS/INS（NAVS）"],
    ["SF.F5",  "应急处置",           "F5: 伞降激活/应急供电切换（PES/EPS）"],
    ["SF.F6",  "提供乘员环境支撑",   "F6: 客舱环境监控与支撑功能"],
    ["SF.F7",  "提供飞行组界面显示", "F7: 状态/告警/视频等显示（IMS/AVCS）"],
    ["SF.F8",  "提供通信",           "F8: DLS链路管理/音视频通信（AVCS）"],
    ["SF.F9",  "提供电能",           "F9: 供配电管理（ES/PACKS/HVDS）"],
    ["SF.F10", "提供数据记录功能",   "F10: 飞行日志/运行数据记录"],
    ["SF.F11", "提供维护功能",       "F11: 地面维护/软件加载/参数配置"],
    ["SF.F12", "提供安全保障",       "F12: 系统安全保障功能"],
    ["SF.F13", "提供探测与避让功能", "F13: DAA 障碍物探测与主动避让（DAAS）"],
    [STOP, "", ""],
]
add_rows(ws4, rows_fn, len(FN_H))
freeze(ws4)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 6: 支持资产  (BDF 的 producer/consumer 实体, 名称须与 BDF 中完全一致)
# 安全域/关键性 ← vol1 表5-3(信任域) + 表5-1(飞行安全关键性)
# 机载核心信任域(D.5)=Internal/High; 机载边界域(D.3)=Shared; 地面/外部=External
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("支持资产")
SA_H = ["编号", "名称", "交联接口", "安全域", "关键性"]
SA_W = [10, 22, 30, 12, 10]
apply_hdr(ws5, SA_H, SA_W, "FF975A16")

dv5a = DataValidation(type="list", formula1='"Internal,External,DMZ,Shared"', allow_blank=True)
ws5.add_data_validation(dv5a)
dv5a.sqref = "D2:D100"
dv5b = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws5.add_data_validation(dv5b)
dv5b.sqref = "E2:E100"

# subsystem → (security_domain, criticality)  per vol1 表5-3 / 表5-1
ONBOARD_CORE = {"ICP", "EPS", "PACKS", "HVDS", "NAVS", "DAAS", "PES"}  # 机载核心信任域 → Internal/High


def classify(name):
    if name in ONBOARD_CORE:
        return ("Internal", "High")
    if name == "DLS":
        return ("Shared", "High")    # 空地数据链骨干 (C2)
    if name == "AVCS":
        return ("Shared", "Medium")  # 音视频通信 (支撑)
    return ("External", "Medium")    # 地面/外部实体


entities = []
for _, _, prod, cons, *_ in bdf_rows:
    for n in (prod, cons):
        if n not in entities:
            entities.append(n)
rows_sa = []
for i, name in enumerate(entities, 1):
    dom, crit = classify(name)
    rows_sa.append([f"ASA.{i:02d}", name, "", dom, crit])
rows_sa.append([STOP, "", "", "", ""])
add_rows(ws5, rows_sa, len(SA_H))
freeze(ws5)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 7: STRIDE映射表  (reference only)
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("STRIDE映射表")
ST_H = ["数据流类型", "主要STRIDE类别", "典型攻击模式", "安保措施方向", "FHA严重度参考"]
ST_W = [14, 26, 42, 48, 20]
apply_hdr(ws6, ST_H, ST_W, "FF9B2C2C")

stride_data = [
    ["CMD", "Spoofing + Tampering", "未授权指令注入/篡改/重放；推力/上下电/伞降误触发",
     "双向认证 + 完整性保护(MAC/签名) + 序列号/时间窗 + 运行状态门控", "Catastrophic / Hazardous"],
    ["CONFIG", "Tampering", "配置数据恶意修改(航线/围栏)；未授权维护参数篡改",
     "签名 + 语义约束检查 + 版本控制 + 强认证/RBAC", "Hazardous / Major"],
    ["LOAD", "Tampering", "恶意软件加载(维护接口)；未授权维护访问",
     "Secure Boot + 软件签名验证 + 完整性链 + 强认证/RBAC", "Catastrophic"],
    ["STATE", "Spoofing", "状态/遥测伪造；导航融合异常",
     "多源融合一致性校验 + 异常检测 + 冗余", "Major / Hazardous"],
    ["ALERT", "Tampering + DenialOfService", "告警抑制/篡改；总线阻塞DoS",
     "完整性保护 + 实时调度/带宽隔离 + 告警超时检测", "Hazardous / Catastrophic"],
    ["DATA", "InformationDisclosure", "状态/视频数据泄露；日志导出暴露",
     "加密传输(TLS) + 访问控制 + 流媒体加密", "Minor / Major"],
    ["SENSOR", "Spoofing + Tampering", "GNSS欺骗/干扰；探测/感知输入伪造",
     "多源一致性 + 物理合理性约束 + 冗余传感器交叉验证", "Major / Hazardous"],
]
bgs = ["FFFFF5F5", "FFF0FFF4", "FFFFF5F5", "FFEBF8FF", "FFFFF5F5", "FFF0FFF4", "FFEBF8FF"]
for ri, (row, bg) in enumerate(zip(stride_data, bgs), 2):
    for ci, v in enumerate(row, 1):
        c = ws6.cell(row=ri, column=ci, value=v)
        c.font = norm()
        c.alignment = lft()
        c.border = bdr()
        c.fill = fill(bg)
    ws6.row_dimensions[ri].height = 44
ws6.row_dimensions[1].height = 28
ws6.freeze_panes = "A2"

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 8: 信任边界表  ←  vol1 表5-4 / vol3 表4-1
# ─────────────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("信任边界表")
TB_H = ["边界编号", "边界名称", "边界说明", "覆盖域", "相关威胁主体", "是否进入内部传播"]
TB_W = [12, 24, 46, 18, 28, 16]
apply_hdr(ws7, TB_H, TB_W, "FF0E7490")

dv7 = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
ws7.add_data_validation(dv7)
dv7.sqref = "F2:F50"

rows_tb = [
    ["SB-01", "外部网络与通信边界", "系统与外部网络、云平台、通信基础设施及操控终端之间的数据交换边界。",
     "D.1;D.2;D.3", "TA-E-01;TA-E-02;TA-I-01", "是"],
    ["SB-02", "开放环境输入边界", "系统接收来自开放环境的导航、感知和探测信息的边界。",
     "D.1;D.4", "TA-E-03;TA-E-04", "是"],
    ["SB-03", "维护与物理接入边界", "系统软件加载、参数配置、调试维护和设备接入边界。",
     "D.6;D.5", "TA-E-05;TA-I-02;TA-I-03;TA-I-04", "是"],
    [STOP, "", "", "", "", ""],
]
add_rows(ws7, rows_tb, len(TB_H))
freeze(ws7)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 9: 威胁主体表  ←  vol1 表6-5/6-6
# ─────────────────────────────────────────────────────────────────────────────
ws8 = wb.create_sheet("威胁主体表")
TA_H = ["主体编号", "名称", "类型", "说明"]
TA_W = [12, 24, 12, 50]
apply_hdr(ws8, TA_H, TA_W, "FFB91C1C")

dv8 = DataValidation(type="list", formula1='"外部,内部,第三方"', allow_blank=True)
ws8.add_data_validation(dv8)
dv8.sqref = "C2:C50"

rows_ta = [
    ["TA-E-01", "远程网络攻击者",     "外部",   "经公网/运营商网络对空地数据链及云平台发起远程攻击。"],
    ["TA-E-02", "通信链路攻击者",     "外部",   "针对1.4G/4G/5G射频链路的中间人/重放/注入攻击。"],
    ["TA-E-03", "GNSS欺骗/干扰者",    "外部",   "对GNSS卫星信号进行欺骗或干扰，影响导航定位。"],
    ["TA-E-04", "开放环境输入伪造者", "外部",   "伪造探测/感知/导航等开放环境输入数据。"],
    ["TA-E-05", "恶意维护人员/设备",  "外部",   "通过维护接口加载恶意软件或篡改参数。"],
    ["TA-I-01", "受感染地面终端",     "内部",   "被攻陷的RCS/转发服务器成为内部攻击跳板。"],
    ["TA-I-02", "维护误操作",         "内部",   "维护人员误操作导致配置或软件异常。"],
    ["TA-I-03", "供应链植入",         "第三方", "供应链环节植入的后门或被篡改的组件。"],
    ["TA-I-04", "内部越权访问",       "内部",   "内部人员越权访问或滥用维护/管理权限。"],
    [STOP, "", "", ""],
]
add_rows(ws8, rows_ta, len(TA_H))
freeze(ws8)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 10: 系统数据流表  ←  vol3 附录E (SDF01-SDF55, 航空器内部 producer→consumer)
# cols: 数据流编号 | 产生者 | 用户 | 内容 | 数据流类型 | 目标功能
# ─────────────────────────────────────────────────────────────────────────────
ws9 = wb.create_sheet("系统数据流表")
SDF_H = ["数据流编号", "产生者", "用户", "内容", "数据流类型", "目标功能"]
SDF_W = [12, 10, 10, 50, 12, 16]
apply_hdr(ws9, SDF_H, SDF_W, "FF0F766E")

dv9 = DataValidation(type="list", formula1='"CMD,CONFIG,STATE,DATA,LOAD,ALERT,SENSOR"', allow_blank=True, showDropDown=False)
ws9.add_data_validation(dv9)
dv9.sqref = "E2:E200"

# (sdf, producer, consumer, content, type, raw_func)
sdf_rows = [
    ("SDF01", "RCS", "DLS", "电子围栏、航线、备降点信息", "CONFIG", "F3"),
    ("SDF02", "RCS", "DLS", "起飞、悬停、恢复任务、备降、上下电、低压配电盒通道开关及复位、电源变换器复位等控制指令", "CMD", "F5"),
    ("SDF03", "RCS", "DLS", "飞行机组语音", "DATA", "F8"),
    ("SDF04", "DLS", "RCS", "飞机运行状态", "STATE", "F7"),
    ("SDF05", "DLS", "RCS", "飞机告警信息", "ALERT", "F5,F7"),
    ("SDF06", "DLS", "RCS", "音视频数据", "DATA", "F7"),
    ("SDF07", "DLS", "IMS", "电子围栏、航线、备降点信息", "CONFIG", "F3"),
    ("SDF08", "DLS", "IMS", "起飞、悬停、恢复任务、备降、上下电、低压配电盒通道开关及复位、电源变换器复位等控制指令", "CMD", "F5"),
    ("SDF09", "IMS", "DLS", "飞机运行状态", "STATE", "F7"),
    ("SDF10", "IMS", "DLS", "飞机告警信息", "ALERT", "F5,F7"),
    ("SDF11", "AVCS", "DLS", "乘客语音", "DATA", "F8"),
    ("SDF12", "AVCS", "DLS", "舱内摄像头视频流", "DATA", "F7"),
    ("SDF13", "AVCS", "DLS", "下视摄像头视频流", "DATA", "F7"),
    ("SDF14", "DLS", "AVCS", "飞行机组语音", "DATA", "F8"),
    ("SDF15", "DAAS", "DLS", "叠加探测信息的视频流", "DATA", "F7"),
    ("SDF16", "DAAS", "IMS", "障碍物类型、距离等状态", "STATE", "F13"),
    ("SDF17", "DAAS", "IMS", "障碍物类型、距离等告警；刹停告警建议", "ALERT", "F13"),
    ("SDF18", "IMS", "DAAS", "姿态角、飞行速度等数据", "STATE", "F4"),
    ("SDF19", "EPS", "IMS", "转速等状态信息", "STATE", "F1,F2,F3,F7"),
    ("SDF20", "EPS", "IMS", "温度、电压、电流等状态信息", "STATE", "F1,F3,F7"),
    ("SDF21", "EPS", "IMS", "转速、温度、电压、电流、故障代码等告警信息", "ALERT", "F1,F2,F7"),
    ("SDF22", "IMS", "EPS", "转速控制指令", "CMD", "F1,F2"),
    ("SDF23", "ES", "IMS", "电源变换器、低压配电盒、应急配电盒、应急电池状态信息", "STATE", "F2,F3,F7,F9"),
    ("SDF24", "ES", "IMS", "电源变换器、低压配电盒、应急配电盒、应急电池告警信息", "ALERT", "F2,F7,F9"),
    ("SDF25", "IMS", "ES", "复位/通道通断指令", "CMD", "F2,F9"),
    ("SDF26", "HVDS", "IMS", "高压配电盒状态信息", "STATE", "F1,F2,F3,F7,F9"),
    ("SDF27", "HVDS", "IMS", "高压配电盒告警信息", "ALERT", "F1,F2,F7,F9"),
    ("SDF28", "IMS", "HVDS", "空地状态、电机电压、动力电池电压电流值等信息", "STATE", "F1,F9"),
    ("SDF29", "PACKS", "IMS", "电池电压、电流、温度等状态信息", "STATE", "F1,F2,F3,F7,F9"),
    ("SDF30", "PACKS", "IMS", "电池电压、电流、温度等告警信息", "ALERT", "F1,F2,F7,F9"),
    ("SDF31", "IMS", "PACKS", "高压上下电指令", "CMD", "F1,F2,F3,F9"),
    ("SDF32", "IMS", "PACKS", "航空器空地状态、高压配电盒支路继电器状态", "STATE", "F1,F9"),
    ("SDF33", "NAVS", "IMS", "导航计算机、卫星接收机、卫星天线、三轴陀螺、三轴加速度计等状态信息", "STATE", "F2,F3,F4,F7"),
    ("SDF34", "NAVS", "IMS", "航向、姿态、速度、加速度、位置等数据信息", "STATE", "F2,F3,F4,F7"),
    ("SDF35", "NAVS", "IMS", "航向、姿态、速度、加速度、GNSS位置等告警信息", "ALERT", "F2,F3,F4,F7"),
    ("SDF36", "IMS", "NAVS", "RTCM纠偏数据", "DATA", "F2,F3,F4,F7"),
    ("SDF37", "PES", "IMS", "状态信息", "STATE", "F5,F7"),
    ("SDF38", "PES", "IMS", "告警信息", "ALERT", "F5,F7"),
    ("SDF39", "IMS", "PES", "开伞指令", "CMD", "F5,F2"),
    ("SDF40", "IMS", "PES", "加解锁指令", "CMD", "F5"),
    ("SDF41", "IMS", "FCS", "导航数据", "CONFIG", "F2,F3,F4,F7"),
    ("SDF42", "IMS", "FCS", "备降点数据", "CONFIG", "F2,F3,F5,F7"),
    ("SDF43", "IMS", "FCS", "电子围栏数据", "CONFIG", "F2,F3,F7"),
    ("SDF44", "IMS", "FCS", "应急降落微调指令数据", "CMD", "F2,F3,F5,F7"),
    ("SDF45", "FCS", "IMS", "转速控制指令", "CMD", "F1,F2,F7"),
    ("SDF46", "FCS", "IMS", "内外环控制量，当前飞行状态及模式", "STATE", "F2,F7"),
    ("SDF47", "FCS", "IMS", "飞行包线告警信息，电子围栏告警信息", "ALERT", "F2,F3,F5,F7"),
    ("SDF48", "IMS", "FMS", "导航数据", "STATE", "F3,F4,F7"),
    ("SDF49", "IMS", "FMS", "探测与避让系统数据", "STATE", "F3,F7,F13"),
    ("SDF50", "IMS", "FMS", "飞行航线数据", "CONFIG", "F3,F7"),
    ("SDF51", "IMS", "FMS", "动力电池电量、功率数据", "STATE", "F1,F3,F7"),
    ("SDF52", "FMS", "IMS", "待飞时间和待飞航程，剩余续航时间和续航里程", "STATE", "F2,F3,F5,F7"),
    ("SDF53", "FMS", "IMS", "当前飞行状态", "STATE", "F2,F3,F5,F7"),
    ("SDF54", "FMS", "IMS", "避让告警信息", "ALERT", "F2,F3,F5,F7,F13"),
    ("SDF55", "FMS", "FCS", "制导指令", "CMD", "F2,F3,F7"),
]
rows_sdf = [[sdf, prod, cons, content, dft, clean_funcs(rawf)] for sdf, prod, cons, content, dft, rawf in sdf_rows]
rows_sdf.append([STOP] + [""] * (len(SDF_H) - 1))
add_rows(ws9, rows_sdf, len(SDF_H))
freeze(ws9)

# ─────────────────────────────────────────────────────────────────────────────
# Finalize
# 注: 功能传播路径(FP, 表4-5)不再手填 —— 由后端 FpAnalysisService 从 BDF+SDF 自动归并生成。
# ─────────────────────────────────────────────────────────────────────────────
wb._sheets = [ws1, ws_bi, ws2, ws3, ws4, ws5, ws6, ws7, ws8, ws9]

out = r"E:/document/airness/project/Attackgraph/docs/资产清单_v4_可导入.xlsx"
wb.save(out)
print("Saved:", out)
print(f"  边界数据流表(BDF) rows: {len(rows_if) - 1}")
print(f"  边界接口表(BI) rows: {len(rows_bi) - 1}")
print(f"  数据资产 rows: {len(rows_da) - 1}")
print(f"  域属性表 rows: {len(rows_dm) - 1}")
print(f"  功能资产 rows: {len(rows_fn) - 1}")
print(f"  支持资产 rows: {len(rows_sa) - 1}")
print(f"  信任边界表 rows: {len(rows_tb) - 1}")
print(f"  威胁主体表 rows: {len(rows_ta) - 1}")
print(f"  系统数据流表(SDF) rows: {len(rows_sdf) - 1}")
